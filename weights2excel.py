from __future__ import division

import torch 
import torch.nn as nn
import torch.nn.functional as F 
from torch.autograd import Variable
import numpy as np
import cv2 
import matplotlib.pyplot as plt
import math
import os

import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, colors, PatternFill
from openpyxl.chart import BarChart, Reference, label, LineChart

from util import count_parameters as count
from util import convert2cpu as cpu
from util import predict_transform

def print_log(text="", fp_log=None, level=0):
    #torch.set_printoptions(precision=10)

    text = str(text)

    for i in range(level):
        text = "    "+text

    if (fp_log):
        print(text, file=fp_log)

    print(text)

def tensor_equal(tensor1, tensor2, cmp_precision=4):
    no_diff = True;

    tensor_delta = torch.abs(tensor2 - tensor1)
    if torch.equal(tensor1, tensor2) == False:
        no_diff = False
        if cmp_precision > 0:
            diff_delta = 1
            for i in range(cmp_precision):
               diff_delta /= 10
            tensor_delta = torch.gt(tensor_delta, diff_delta)

    tensor_index = torch.nonzero(tensor_delta.view(tensor_delta.numel()))
    if tensor_index.numel() == 0:
        no_diff = True

    return no_diff, tensor_index.reshape(tensor_index.numel()).tolist()

def tensor_visualize_diff(vis, tensor_weights, title):
    XVal = tensor_weights.view(tensor_weights.numel())
    XVar = torch.var(XVal)
    XMin = torch.min(XVal)
    XMax = torch.max(XVal)
    XMean = torch.mean(XVal)
    if XVar.item() == 0:
        bins = 10
    else:
        bins = round((XMax.item()-XMin.item())/XVar.item())
    if bins < 2:
        bins = 10
    nZero = XVal.numel() - torch.nonzero(XVal).size(0)
    sTitle = title+"[{}]->bins:{} var:{:2.3f} min:{:2.3f} max:{:2.3f} mean:{:2.3f} zero:{}".format(XVal.numel(), bins, XVar.item(), XMin.item(), XMax.item(), XMean.item(), nZero)

    vis.histogram(XVal, opts=dict(title=sTitle, numbins=bins))


def is_hot_point(x_val, x_percent):
    val = abs(x_val)
    percent = abs(x_percent)
    if val >= 10. and percent >= 0.01:
        return True
    elif val >= 1. and percent >= 0.02:
        return True
    elif val >= 0.1 and percent >= 0.05:
        return True
    elif val >= 0.01 and percent >= 0.15:
        return True
    elif val >= 0.001 and percent >= 0.25:
        return True
    elif val >= 0.0001 and percent >= 0.50:
        return True
    elif  percent >= 1.:
        return True
    return False

def tensor_diff_evaluation(vis, tensor_ref, tensor_delta_percent, tensor_diff_index, title="", log=None, level=0, max_index=10000):
    hot_point = []
    x_ref = tensor_ref.view(tensor_ref.numel())
    x_percent = tensor_delta_percent.view(tensor_delta_percent.numel())
    index_range = x_ref.numel()
    if index_range > max_index:
        index_range = max_index
    for i in range(index_range):
        if i in tensor_diff_index and is_hot_point(x_ref[i], x_percent[i]):
            hot_point.append(i)

    if 0 == len(hot_point):
        print_log("++  hot    "+str(len(hot_point))+", diff can be ignored", log, level)
    else:
        print_log("++  hot    "+str(len(hot_point))+", index "+str(hot_point), log, level)

    return hot_point


def tensor_diff(vis, tensor_cmp, tensor_ref, cmp_precision=4, title="", log=None, level=0):
    is_acceptable = False
    no_diff, diff_index = tensor_equal(tensor_ref, tensor_cmp, cmp_precision)
    if no_diff == False:
        tensor_ref = tensor_ref.view(tensor_ref.numel())
        tensor_cmp = tensor_cmp.view(tensor_cmp.numel())
        tensor_delta = tensor_cmp - tensor_ref
        tensor_delta_percent = tensor_delta / tensor_ref
        print_log("--- "+title+" mismatch ---", log, level)
        print_log("++ "+str(tensor_ref.size()), log, level+1)
        print_log("++ golden  "+str(tensor_ref), log, level+1)
        print_log("++  yours  "+str(tensor_cmp), log, level+1)
        print_log("++  delta  "+str(tensor_delta), log, level+1)
        print_log("++  delta% "+str(tensor_delta_percent), log, level+1)
        print_log("++  diffs  "+str(len(diff_index))+", index "+str(diff_index), log, level+1)

        hot_point = tensor_diff_evaluation(vis, tensor_ref, tensor_delta_percent, diff_index, title, log, level+1)
        if 0 == len(hot_point):
            is_acceptable = True

        #vis
        if vis is not None:
            tensor_visualize_diff(vis, tensor_ref, "ref "+title)
            tensor_visualize_diff(vis, tensor_cmp, "your "+title)
            tensor_visualize_diff(vis, tensor_delta, "delta "+title)
            #tensor_visualize_diff(vis, tensor_delta_percent, "delta% "+title)

        #slice diff
        #if title.endswith("conv weight"):
        #    x = tensor_ref
        #    x = x.reshape(x.shape[0], x.shape[1]*x.shape[2]*x.shape[3])
        #    x_cmp = tensor_cmp
        #    x_cmp = x_cmp.reshape(x_cmp.shape[0], x_cmp.shape[1]*x_cmp.shape[2]*x_cmp.shape[3])
        #    for i in range(x.shape[0]):
        #        sSlice = "slice [{}]".format(i)
        #        tensor_diff(vis, x[i], x_cmp[i], cmp_precision, sSlice, log, level+1)
        #    print_log(fp_log=log)
    else:
        print_log("--- "+title+" no diff ---", log, level)

    print_log(fp_log=log)

    return no_diff, is_acceptable


class test_net(nn.Module):
    def __init__(self, num_layers, input_size):
        super(test_net, self).__init__()
        self.num_layers= num_layers
        self.linear_1 = nn.Linear(input_size, 5)
        self.middle = nn.ModuleList([nn.Linear(5,5) for x in range(num_layers)])
        self.output = nn.Linear(5,2)
    
    def forward(self, x):
        x = x.view(-1)
        fwd = nn.Sequential(self.linear_1, *self.middle, self.output)
        return fwd(x)
        
def rgb_images_of_weights(weights):
    """
    Convert weights into a bunch of rgb images. 1 maps to red, -1 maps to blue, 0 to green. Others values are mixed.
    
    Args:
        weights: torch.Tensor with the shape of [n, width, height]
        
    Returns:
        images: torch.Tensor with shape of [n, 3, width, height]
    """
    weights = torch.unsqueeze(weights, 1)
    r = torch.clamp(weights, min=0.0)
    b = torch.clamp(-weights, min=0.0)
    g = 1 - r - b
    img = torch.cat([(r - 0.5) * 2, (g - 0.5) * 2, (b - 0.5) * 2], 1)
    return img

def letterbox_image(img, inp_dim):
    '''resize image with unchanged aspect ratio using padding'''
    img_w, img_h = img.shape[1], img.shape[0]
    w, h = inp_dim
    new_w = int(img_w * min(w/img_w, h/img_h))
    new_h = int(img_h * min(w/img_w, h/img_h))
    resized_image = cv2.resize(img, (new_w,new_h), interpolation = cv2.INTER_CUBIC)

    canvas = np.full((inp_dim[1], inp_dim[0], 3), 128)

    canvas[(h-new_h)//2:(h-new_h)//2 + new_h,(w-new_w)//2:(w-new_w)//2 + new_w,  :] = resized_image

    return canvas

def get_test_input(input_dim, CUDA, imgFile="/AIHOME/proj/pyWeightAna/dog-cycle-car.png", ratio=True, debug=False):
    img = cv2.imread(imgFile)
    if ratio:
        img = letterbox_image(img, (input_dim, input_dim))
    else:
        img = cv2.resize(img, (input_dim, input_dim)) 
    img_ =  img[:,:,::-1].transpose((2,0,1))
    img_ = img_[np.newaxis,:,:,:]/255.0
    img_ = torch.from_numpy(img_).float()
    img_ = Variable(img_)
    if CUDA:
        img_ = img_.cuda()
#    if debug:
#        print(img_.size())
    return img_


def parse_cfg(cfgfile):
    """
    Takes a configuration file
    
    Returns a list of blocks. Each blocks describes a block in the neural
    network to be built. Block is represented as a dictionary in the list
    
    """
    file = open(cfgfile, 'r')
    lines = file.read().split('\n')     #store the lines in a list
    lines = [x for x in lines if len(x) > 0] #get read of the empty lines 
    lines = [x for x in lines if x[0] != '#']  
    lines = [x.rstrip().lstrip() for x in lines]

    
    block = {}
    blocks = []
    
    for line in lines:
        if line[0] == "[":               #This marks the start of a new block
            if len(block) != 0:
                blocks.append(block)
                block = {}
            block["type"] = line[1:-1].rstrip()
        else:
            key,value = line.split("=")
            block[key.rstrip()] = value.lstrip()
    blocks.append(block)

    return blocks
#    print('\n\n'.join([repr(x) for x in blocks]))

import pickle as pkl

class MaxPoolStride1(nn.Module):
    def __init__(self, kernel_size):
        super(MaxPoolStride1, self).__init__()
        self.kernel_size = kernel_size
        self.pad = kernel_size - 1
    
    def forward(self, x):
        padded_x = F.pad(x, (0,self.pad,0,self.pad), mode="replicate")
        pooled_x = nn.MaxPool2d(self.kernel_size, self.pad)(padded_x)
        return pooled_x
    

class EmptyLayer(nn.Module):
    def __init__(self):
        super(EmptyLayer, self).__init__()
        

class DetectionLayer(nn.Module):
    def __init__(self, anchors):
        super(DetectionLayer, self).__init__()
        self.anchors = anchors
    
    def forward(self, x, inp_dim, num_classes, confidence):
        x = x.data
        global CUDA
        prediction = x
        prediction = predict_transform(prediction, inp_dim, self.anchors, num_classes, confidence, CUDA)
        return prediction
        

        


class Upsample(nn.Module):
    def __init__(self, stride=2):
        super(Upsample, self).__init__()
        self.stride = stride
        
    def forward(self, x):
        stride = self.stride
        assert(x.data.dim() == 4)
        B = x.data.size(0)
        C = x.data.size(1)
        H = x.data.size(2)
        W = x.data.size(3)
        ws = stride
        hs = stride
        x = x.view(B, C, H, 1, W, 1).expand(B, C, H, stride, W, stride).contiguous().view(B, C, H*stride, W*stride)
        return x
#       
        
class ReOrgLayer(nn.Module):
    def __init__(self, stride = 2):
        super(ReOrgLayer, self).__init__()
        self.stride= stride
        
    def forward(self,x):
        assert(x.data.dim() == 4)
        B,C,H,W = x.data.shape
        hs = self.stride
        ws = self.stride
        assert(H % hs == 0),  "The stride " + str(self.stride) + " is not a proper divisor of height " + str(H)
        assert(W % ws == 0),  "The stride " + str(self.stride) + " is not a proper divisor of height " + str(W)
        x = x.view(B,C, H // hs, hs, W // ws, ws).transpose(-2,-3).contiguous()
        x = x.view(B,C, H // hs * W // ws, hs, ws)
        x = x.view(B,C, H // hs * W // ws, hs*ws).transpose(-1,-2).contiguous()
        x = x.view(B, C, ws*hs, H // ws, W // ws).transpose(1,2).contiguous()
        x = x.view(B, C*ws*hs, H // ws, W // ws)
        return x


def create_modules(blocks, print_net=True):
    net_info = blocks[0]     #Captures the information about the input and pre-processing    
    
    module_list = nn.ModuleList()
    
    index = 0    #indexing blocks helps with implementing route  layers (skip connections)

    
    prev_filters = 3
    
    output_filters = []
    
    for x in blocks:
        if (x["type"] == "net"):
            continue

        if (print_net is True):
            print(str(index)+" "+str(x))

        module = nn.Sequential()
        
        #fx YOLO input layer
        if (x["type"] == "input"):
            module_list.append(module)
            output_filters.append(0)
            index += 1
            continue
        
        #If it's a convolutional layer
        if (x["type"] == "convolutional" or x["type"] == "dilated_convolutional"):#dilated_convolutional for deeplab
            #Get the info about the layer
            try:
                activation = x["activation"]
            except:
                activation = ""

            try:
                batch_normalize = int(x["batch_normalize"])
            except:
                batch_normalize = 0

            if batch_normalize > 0:
                bias = False
            else:
                bias = True
                
            filters= int(x["filters"])
            padding = int(x.get("pad", 0))
            try:
                kernel_size = int(x["size"])
                kernel_size_w = kernel_size
                kernel_size_h = kernel_size
            except:
                kernel_size_w =  int(x["size_w"])
                kernel_size_h =  int(x["size_h"])
            stride = int(x["stride"])
            rate = int(x.get("rate", 1))
            groups = int(x.get("groups", 1))
            
            if padding:
                pad = (kernel_size - 1) // 2
            else:
                pad = 0
                
            #Add the convolutional layer
            #conv = nn.Conv2d(prev_filters, filters, kernel_size, stride, pad, bias = bias, dilation=rate)
            conv = nn.Conv2d(prev_filters, filters, (kernel_size_h, kernel_size_w), stride, pad, bias = bias, dilation=rate, groups=groups)
            if rate > 1:
                module.add_module("dilated_conv_{0}".format(index), conv)
            else:
                module.add_module("conv_{0}".format(index), conv)
            
            #Add the Batch Norm Layer
            if batch_normalize:
                bn = nn.BatchNorm2d(filters)
                module.add_module("batch_norm_{0}".format(index), bn)
            
            #Check the activation. 
            #It is either Linear or a Leaky ReLU for YOLO
            if activation == "leaky":
                activn = nn.LeakyReLU(0.1, inplace = True)
                module.add_module("leaky_{0}".format(index), activn)
            
            
            
        #If it's an upsampling layer
        #We use Bilinear2dUpsampling
        
        elif (x["type"] == "upsample"):
            stride = int(x["stride"])
#            upsample = Upsample(stride)
            upsample = nn.Upsample(scale_factor = 2, mode = "nearest")
            module.add_module("upsample_{}".format(index), upsample)
        
        #If it is a route layer
        elif (x["type"] == "route"):
            route = EmptyLayer()
            module.add_module("route_{0}".format(index), route)
            
            x["layers"] = x["layers"].split(',')
            
            filters = 0
            for layer in x["layers"]:
                i = int(layer)
                if i > 0:
                    i -= index
                if i < 0:
                    filters += output_filters[index + i]

            ##Start  of a route
            #start = int(x["layers"][0])
            #
            ##end, if there exists one.
            #try:
            #    end = int(x["layers"][1])
            #except:
            #    end = 0
            #
            ##Positive anotation
            #if start > 0:
            #    start = start - index
            #
            #if end > 0:
            #    end = end - index
            #
            #if end < 0:
            #    filters = output_filters[index + start] + output_filters[index + end]
            #else:
            #    filters= output_filters[index + start]
            
        
        #shortcut corresponds to skip connection
        elif x["type"] == "shortcut":
            from_ = int(x["from"])
            shortcut = EmptyLayer()
            module.add_module("shortcut_{}".format(index), shortcut)
            
            
        elif x["type"] == "maxpool":
            stride = int(x["stride"])
            size = int(x["size"])
            if stride != 1:
                maxpool = nn.MaxPool2d(size, stride)
            else:
                maxpool = MaxPoolStride1(size)
            
            module.add_module("maxpool_{}".format(index), maxpool)
        
        #Yolo is the detection layer
        elif x["type"] == "yolo":
            mask = x["mask"].split(",")
            mask = [int(x) for x in mask]
            
            
            anchors = x["anchors"].split(",")
            anchors = [int(a) for a in anchors]
            anchors = [(anchors[i], anchors[i+1]) for i in range(0, len(anchors),2)]
            anchors = [anchors[i] for i in mask]
            
            detection = DetectionLayer(anchors)
            module.add_module("Detection_{}".format(index), detection)
        
        elif x["type"] == "softmax":#for deeplab    
            spatial = int(x["spatial"])
            softmax = EmptyLayer()
            module.add_module("softmax_{}".format(index), softmax)
        
        elif x["type"] == "bilinear":#for deeplab    
            stride = int(x["stride"])
            bilinear = nn.UpsamplingBilinear2d(scale_factor=stride);
            module.add_module("bilinear_{}".format(index), bilinear)
        
        elif x["type"] == "avgpool":#for deeplab    
            avgpool = nn.AvgPool2d(1);
            module.add_module("avgpool_{}".format(index), avgpool)

        ####
        #elif x["type"] == "connected":#for vgg16
        #    output = int(x["output"])
        #    connected  = nn.Linear( ,output)

        elif x["type"] == "dropout":#for vgg16
            dropout  = nn.Dropout(p=0.5);
            module.add_module("dropout_{}".format(index), dropout)

        else:
            print("Something I dunno: "+x["type"])
            assert False


        module_list.append(module)
        prev_filters = filters
        output_filters.append(filters)
        index += 1
        
    
    return (net_info, module_list)



class Darknet(nn.Module):
    def __init__(self, cfgfile, print_net=True):
        super(Darknet, self).__init__()
        self.blocks = parse_cfg(cfgfile)
        self.net_info, self.module_list = create_modules(self.blocks, print_net)
        self.header = torch.IntTensor([0,2,0,0])
        self.seen = 0
        self.result = None
   
     
    def get_blocks(self):
        return self.blocks
    
    def get_module_list(self):
        return self.module_list

                
    def forward(self, x, CUDA, vis=None, featureLayers=None, show=3, view='flat', conf=0.5):
        detections = []
        modules = self.blocks[1:]
        outputs = {}   #We cache the outputs for the route layer
        
        
        write = 0
        for i in range(len(modules)):        
            
            module_type = (modules[i]["type"])
            if module_type == "convolutional" or module_type == "dilated_convolutional" or module_type == "upsample" or module_type == "maxpool" or module_type == "bilinear" or module_type == "avgpool":
                
                x = self.module_list[i](x)
                outputs[i] = x

                #visualize feature map
                if vis and module_type == "convolutional" and (featureLayers==None or len(featureLayers)==0 or featureLayers.split().count(str(i+1))>0):
                    self.visualizeFeatureMap(vis, i, x, show, view)
                
            elif module_type == "softmax":
                return F.softmax(x);

            elif module_type == "route":
                layers = modules[i]["layers"]
                layers = [int(a) for a in layers]
                
                if (layers[0]) > 0:
                    layers[0] = layers[0] - i

                if len(layers) == 1:
                    x = outputs[i + (layers[0])]

                else:
                    if (layers[1]) > 0:
                        layers[1] = layers[1] - i
                        
                    map1 = outputs[i + layers[0]]
                    map2 = outputs[i + layers[1]]
                    
                    
                    x = torch.cat((map1, map2), 1)
                outputs[i] = x
            
            elif  module_type == "shortcut":
                from_ = int(modules[i]["from"])
                x = outputs[i-1] + outputs[i+from_]
                outputs[i] = x
                
            
            
            elif module_type == 'yolo':        
                
                anchors = self.module_list[i][0].anchors
                #Get the input dimensions
                inp_dim = int (self.net_info["height"])
                
                #Get the number of classes
                num_classes = int (modules[i]["classes"])
                
                grid = x.size(2)
                #Output the result
                x = x.data
                x = predict_transform(x, inp_dim, anchors, num_classes, CUDA)
                
                if type(x) == int:
                    continue

                #visualize feature map
                if vis and (featureLayers==None or len(featureLayers)==0 or featureLayers.split().count(str(i+1))>0):
                    #X = x.view(1, len(anchors)*(5+num_classes), grid, grid)
                    X = x
                    self.visualizeFeatureMap(vis, i, X, show, view, conf)
                
                if not write:
                    detections = x
                    write = 1
                
                else:
                    detections = torch.cat((detections, x), 1)
                
                outputs[i] = outputs[i-1]
                self.result = outputs[i]#for debug net structure
                
        
        
        try:
            return detections
        except:
            return 0

            
    def load_weights(self, weightfile, layer_num, channel_bins_pow, \
                    vis=None, featureLayers=None, show=3, view="flat", \
                    dn_ref=None, cmp_precision=4, log=None):

        #Open the weights file
        fp = open(weightfile, "rb")

        #The first 4 values are header information 
        # 1. Major version number
        # 2. Minor Version Number
        # 3. Subversion number 
        # 4. IMages seen 
        version = np.fromfile(fp, count=3, dtype=np.int32)
        version = [int(i) for i in version]
        if version[0]*10+version[1] >=2 and version[0] < 1000 and version[1] < 1000:
            seen = np.fromfile(fp, count=1, dtype=np.int64)
        else:
            seen = np.fromfile(fp, count=1, dtype=np.int32)
        self.header = torch.from_numpy(np.concatenate((version, seen), axis=0))
        self.seen = int(seen)

        #The rest of the values are the weights
        # Let's load them up
        ann_version = False
        ANN_FORMAT = 1<<12
        if (ANN_FORMAT == version):
            ann_version = True
        if ann_version is True:
            n_layers = np.fromfile(fp, dtype = np.int32, count = 1)
            fp.seek(4*n_layers[0], 1)
        else:
            weights = np.fromfile(fp, dtype = np.float32)
        
        show_weights = False
        if vis and dn_ref is None:
            show_weights = True

        cmp_weights = False
        if dn_ref is not None:
            cmp_weights = True

        nLayerCmp = 0
        nErr = 0
        nSkip = 0
        ptr = 0

        conv_num = 0

        for i in range(len(self.module_list)):

            module_type = self.blocks[i + 1]["type"]
            
            if module_type == "convolutional" or module_type == "dilated_convolutional":#dilated_convolutional for deeplab
                if (ann_version is True):
                    fp.seek(4*2, 1)
                #print(str(i)+" "+str(self.blocks[i + 1]))
                s_layer = "layer {} {}: ".format(i, module_type)

                show_layer_weights = False 
                if show_weights and (featureLayers==None or len(featureLayers)==0 or featureLayers.split().count(str(i))>0):
                    show_layer_weights = True

                cmp_layer_weights = False
                if cmp_weights and (featureLayers==None or len(featureLayers)==0 or featureLayers.split().count(str(i))>0):
                    cmp_layer_weights = True

                model = self.module_list[i]

                #compare weights
                model_ref = None
                if cmp_layer_weights:
                    nLayerCmp += 1
                    model_ref = dn_ref.module_list[i]

                try:
                    batch_normalize = int(self.blocks[i+1]["batch_normalize"])
                except:
                    batch_normalize = 0
                
                conv = model[0]

                #compare weights
                conv_ref = None
                if model_ref is not None:
                    conv_ref = model_ref[0]
                
                if (batch_normalize > 0):
                    bn = model[1]

                    #compare weights
                    bn_ref = None
                    if model_ref is not None:
                        bn_ref = model_ref[1]
                    
                    #Get the number of weights of Batch Norm Layer
                    num_bn_biases = bn.bias.numel()
                    
                    #Load the weights
                    if (ann_version is True):
                        fp.seek(4, 1)
                        weights = np.fromfile(fp, dtype = np.float32, count=num_bn_biases)
                        ptr = 0
                    bn_biases = torch.from_numpy(weights[ptr:ptr + num_bn_biases])
                    ptr += num_bn_biases
                    
                    if (ann_version is True):
                        fp.seek(4, 1)
                        weights = np.fromfile(fp, dtype = np.float32, count=num_bn_biases)
                        ptr = 0
                    bn_weights = torch.from_numpy(weights[ptr: ptr + num_bn_biases])
                    ptr  += num_bn_biases
                    
                    if (ann_version is True):
                        fp.seek(4, 1)
                        weights = np.fromfile(fp, dtype = np.float32, count=num_bn_biases)
                        ptr = 0
                    bn_running_mean = torch.from_numpy(weights[ptr: ptr + num_bn_biases])
                    ptr  += num_bn_biases
                    
                    if (ann_version is True):
                        fp.seek(4, 1)
                        weights = np.fromfile(fp, dtype = np.float32, count=num_bn_biases)
                        ptr = 0
                    bn_running_var = torch.from_numpy(weights[ptr: ptr + num_bn_biases])
                    ptr  += num_bn_biases
                    
                    #Cast the loaded weights into dims of model weights. 
                    bn_biases = bn_biases.view_as(bn.bias.data)
                    bn_weights = bn_weights.view_as(bn.weight.data)
                    bn_running_mean = bn_running_mean.view_as(bn.running_mean)
                    bn_running_var = bn_running_var.view_as(bn.running_var)

                    #Copy the data to model
                    bn.bias.data.copy_(bn_biases)
                    bn.weight.data.copy_(bn_weights)
                    bn.running_mean.copy_(bn_running_mean)
                    bn.running_var.copy_(bn_running_var)

                    #compare weights
                    if bn_ref:
                        no_diff, is_skip = tensor_diff(vis, bn.bias, bn_ref.bias, cmp_precision, s_layer+"bn bias", log, 0)
                        if no_diff == False:
                            nErr += 1
                            if is_skip == True:
                                nSkip += 1

                        no_diff, is_skip = tensor_diff(vis, bn.weight, bn_ref.weight, cmp_precision, s_layer+"bn weight", log, 0)
                        if no_diff == False:
                            nErr += 1
                            if is_skip == True:
                                nSkip += 1

                        no_diff, is_skip = tensor_diff(vis, bn.running_mean, bn_ref.running_mean, cmp_precision, s_layer+"bn mean", log, 0)
                        if no_diff == False:
                            nErr += 1
                            if is_skip == True:
                                nSkip += 1

                        no_diff, is_skip = tensor_diff(vis, bn.running_var, bn_ref.running_var, cmp_precision, s_layer+"bn var", log, 0)
                        if no_diff == False:
                            nErr += 1
                            if is_skip == True:
                                nSkip += 1
                        #if tensor_diff(vis, bn.bias, bn_ref.bias, cmp_precision, s_layer+"bn bias", log, 0) == False:
                        #    nErr += 1
                        #if tensor_diff(vis, bn.weight, bn_ref.weight, cmp_precision, s_layer+"bn weight", log, 0) == False:
                        #    nErr += 1
                        #if tensor_diff(vis, bn.running_mean, bn_ref.running_mean, cmp_precision, s_layer+"bn mean", log, 0) == False:
                        #    nErr += 1
                        #if tensor_diff(vis, bn.running_var, bn_ref.running_var, cmp_precision, s_layer+"bn var", log, 0) == False:
                        #    nErr += 1

                    if show_layer_weights == True:
                        self.visualizeWeights(vis, i, bn_biases, show, view,  s_layer+"bn bias", log)
                        self.visualizeWeights(vis, i, bn_weights, show, view,  s_layer+"bn weight", log)
                        self.visualizeWeights(vis, i, bn_running_mean, show, view,  s_layer+"bn mean", log)
                        self.visualizeWeights(vis, i, bn_running_var, show, view,  s_layer+"bn var", log)
                else:
                    #Number of biases
                    num_biases = conv.bias.numel()
                
                    #Load the weights
                    if (ann_version is True):
                        fp.seek(4, 1)
                        weights = np.fromfile(fp, dtype = np.float32, count=num_biases)
                        ptr = 0
                    conv_biases = torch.from_numpy(weights[ptr: ptr + num_biases])
                    ptr = ptr + num_biases
                    
                    #reshape the loaded weights according to the dims of the model weights
                    conv_biases = conv_biases.view_as(conv.bias.data)
                    
                    #Finally copy the data
                    conv.bias.data.copy_(conv_biases)

                    #compare weights
                    if conv_ref:
                        no_diff, is_skip = tensor_diff(vis, conv.bias, conv_ref.bias, cmp_precision, s_layer+"conv bias", log, 0)
                        if no_diff == False:
                            nErr += 1
                            if is_skip == True:
                                nSkip += 1
                        #if tensor_diff(vis, conv.bias, conv_ref.bias, cmp_precision, s_layer+"conv bias", log, 0) == False:
                        #    nErr += 1

                    if show_layer_weights:
                        self.visualizeWeights(vis, i, conv_biases, show, view,  s_layer+"conv bias", log)
                    
                #Let us load the weights for the Convolutional layers
                num_weights = conv.weight.numel()
                
                #Do the same as above for weights
                if (ann_version is True):
                    fp.seek(4, 1)
                    weights = np.fromfile(fp, dtype = np.float32, count=num_weights)
                    ptr = 0
                conv_weights = torch.from_numpy(weights[ptr:ptr+num_weights])
                ptr = ptr + num_weights

                conv_weights = conv_weights.view_as(conv.weight.data)
                
                conv_num += 1

                # print("*****************************************************************")
                # print(conv_num)
                # print(i)  # layer num 0~160, 161 is yolo
                # print(conv_weights.shape[0])

                # e.g. layer = 157, conv No.107 convweights.shape = [1024, 512, 3, 3]
                                        
                if i == layer_num:
                    print("*****************************************************************")
                    print('conv_weights.shape: {}'.format(conv_weights.shape))

                    conv_batches = conv_weights.shape[0]    # 1024
                    conv_channels = conv_weights.shape[1]  # 512
                    conv_heights = conv_weights.shape[2]   # 3
                    conv_widthes = conv_weights.shape[3]    # 3
                    print("valid:", conv_batches, conv_channels, conv_heights, conv_widthes)
                    
                    channel_bins = pow(2, channel_bins_pow)   # value is 16, means group_num = 1024/16 = 64
                    channel_bins_num = conv_batches / channel_bins  # 64
                    print('channel_bins: {}'.format(channel_bins))

                    # create excel 157
                    filefullpath = 'output/layer' + str(i) + '.xlsx'
                    if os.path.exists(filefullpath):
                        os.remove(filefullpath)

                    wb = Workbook()
                    # ws = wb.create_sheet('layer' + str(i), i)
                    ws = wb.active
                    ws.freeze_panes = 'B2'
                    
                    alignment_set = Alignment(horizontal='center', vertical='center')
                    fill_set = PatternFill("solid", fgColor="CCCCCC")
                    border_set = Border(left=Side(style='thin', color=colors.BLACK),
                                        right=Side(style='thin', color=colors.BLACK),
                                        top=Side(style='thin', color=colors.BLACK),
                                        bottom=Side(style='thin', color=colors.BLACK))
                    
                    # write 'Channel\Batch' in (1,1)
                    ws.cell(row=1, column=1, value='Channel\Batch')
                    ws.cell(row=1, column=1).font = Font(bold=True)
                    ws.cell(row=1, column=1).alignment = alignment_set
                    ws.cell(row=1, column=1).fill = fill_set
                    ws.cell(row=1, column=1).border = border_set

                    # write 'Channel_Max' in (1, int(channel_bins_num)+2)
                    ws.cell(row=1, column=int(channel_bins_num)+2, value='Channel_Max')
                    ws.cell(row=1, column=int(channel_bins_num)+2).font = Font(bold=True)
                    ws.cell(row=1, column=int(channel_bins_num)+2).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=1, column=int(channel_bins_num)+2).fill = fill_set
                    ws.cell(row=1, column=int(channel_bins_num)+2).border = border_set
                    
                    # write 'Total_Max' in (conv_channels+2, 1)
                    ws.cell(row=conv_channels+2, column=1, value='Total_Max')
                    ws.cell(row=conv_channels+2, column=1).font = Font(bold=True)
                    ws.cell(row=conv_channels+2, column=1).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=conv_channels+2, column=1).fill = fill_set
                    ws.cell(row=conv_channels+2, column=1).border = border_set

                    # write 'Sparsity' in (conv_channels+3, 1)
                    ws.cell(row=conv_channels+3, column=1, value='Sparsity')
                    ws.cell(row=conv_channels+3, column=1).font = Font(bold=True)
                    ws.cell(row=conv_channels+3, column=1).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=conv_channels+3, column=1).fill = fill_set
                    ws.cell(row=conv_channels+3, column=1).border = border_set

                    # write groups = 64 in row
                    for j in range(1, int(channel_bins_num)+1):
                        ws.cell(row=1, column=j+1, value=j)
                        ws.cell(row=1, column=j+1).font = Font(bold=True)
                        ws.cell(row=1, column=j+1).alignment = Alignment(horizontal='center', vertical='center')
                        ws.cell(row=1, column=j+1).fill = fill_set
                        ws.cell(row=1, column=j+1).border = border_set
                    
                    # write channels = 512 in column
                    for k in range(1, conv_channels+1):
                        ws.cell(row=k+1, column=1, value=k)
                        ws.cell(row=k+1, column=1).font = Font(bold=True)
                        ws.cell(row=k+1, column=1).alignment = Alignment(horizontal='center', vertical='center')
                        ws.cell(row=k+1, column=1).fill = fill_set
                        ws.cell(row=k+1, column=1).border = border_set
                    
                    # slice
                    channel_bin_nozero = []
                    channel_bins_max = []
                    total_max = []
                    nozero_len_total = 0
                    for conv_channel in range(conv_channels):  # conv_channels
                        for conv_batch in range(conv_batches):
                            for conv_height in range(conv_heights):
                                for conv_width in range(conv_widthes):
                                    conv_value = conv_weights[conv_batch, conv_channel, conv_height, conv_width]
                                    if conv_value != 0:
                                        channel_bin_nozero.append(conv_value)

                            if (conv_batch + 1) % channel_bins == 0:
                                channel_bin_nozero_len = len(channel_bin_nozero)  # channel bin nozero number 
                                channel_bins_max.append(channel_bin_nozero_len)   # Channel_Max list
                                channel_bins_max_value = max(channel_bins_max)    # Channel_Max value
                                total_max.append(channel_bins_max_value)          # Total_Max list
                                nozero_len_total += channel_bin_nozero_len

                                ws.cell(row=conv_channel+2, column=((conv_batch + 1) / channel_bins + 1), value=channel_bin_nozero_len)
                                ws.cell(row=conv_channel+2, column=((conv_batch + 1) / channel_bins + 1)).alignment = Alignment(horizontal='center', vertical='center')

                                ws.cell(row=conv_channel+2, column=channel_bins_num+2, value=channel_bins_max_value)
                                ws.cell(row=conv_channel+2, column=channel_bins_num+2).alignment = Alignment(horizontal='center', vertical='center')

                                channel_bin_nozero.clear()
                        channel_bins_max.clear()

                    total_max_value = max(total_max)                  # Total_Max value
                    ws.cell(row=conv_channels+2, column=2, value=total_max_value)
                    ws.cell(row=conv_channels+2, column=2).alignment = Alignment(horizontal='center', vertical='center')

                    sparsity = (conv_batches*conv_channels*conv_widthes*conv_heights - nozero_len_total) / (conv_batches*conv_channels*conv_widthes*conv_heights)
                    sparsity = "%.2f%%" % (sparsity * 100)

                    ws.cell(row=conv_channels+3, column=2, value=sparsity)
                    ws.cell(row=conv_channels+3, column=2).alignment = Alignment(horizontal='center', vertical='center')                    

                    print('nozero_len_total: ', nozero_len_total)
                    # print('sparsity: %.2f%%'%(sparsity*100))
                    print('sparsity: ', sparsity)
                    print('total_max: ', total_max_value)

                    # #create chart
                    # bar_chart = BarChart()
                    # # bar_chart.type = 'col'
                    # bar_chart.title = 'Weights Nozero Distribution'
                    # # bar_chart.style = 12
                    # bar_chart.x_axis.title = 'Channels'
                    # bar_chart.x_axis.delete = False
                    # bar_chart.y_axis.title = 'Channel bins max'
                    # bar_chart.y_axis.delete = False
                    # lables = Reference(ws, min_row=2, max_row=10+1, min_col=1)
                    # data = Reference(ws, min_row=1, max_row=10+1, min_col=channel_bins_num+2)
                    # bar_chart.add_data(data, titles_from_data=True)
                    # bar_chart.set_categories(lables)
                    # # bar_chart.shape = 0
                    # bar_chart.height = 7.5
                    # bar_chart.width = 15
                    # bar_chart.dLbls = label.DataLabelList()
                    # bar_chart.dLbls.showVal = True
                    # bar_chart.showVal = True

                    # ws.add_chart(bar_chart,'B13')

                    wb.save(filefullpath) 
                    
                    # ct_excel=pd.read_excel(filefullpath)
                    # print(ct_excel)
                    
                    # ct_excel.sort_values(by='Channel_Max', inplace=True, ascending=False)
                    # ct_excel.plot.bar(x='Channel_Batch', y='Channel_Max', color='blue', title='Weights Nozero Distribution')   
                    # ct_excel.plot.histogram()
                    # plt.tight_layout() 
                    # plt.show()

                    # plt.savefig( 'output/layer' + str(i) + '.png')


if __name__ == "__main__":
    '''
    e.g. layer_num = 157: conv.shape=[1024, 512, 3, 3]
    channel_bins_pow=4 means 2**4 = 16
    group = 1024/step = 64

    layer_num = 158 conv.shape=[512, 1024, 1, 1]

    layer_num = 148 conv.shape=[512, 256, 3, 3]

    '''
    net_cfg = Darknet('yolov4.cfg')
    net_cfg.load_weights("yolov4.best.0", layer_num=148, channel_bins_pow=4)  



